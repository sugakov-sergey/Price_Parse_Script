"""
Объединяет все файлы из папки 'input-for-merge' в один
файл 'merged_price'. Условие объединения при совпадении
позиций указываются до начала обработки в консоли.
"""
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import BUILTIN_FORMATS

from start_app import _cell_alignment, _make_table_view

print("Внимание!\n"
      "Для правильного объедиения файлов прайсы должны быть обработаны \n"
      "и сохранены в папке 'input-for-merge'\n")

msg = 'Какой режим обработки прайса ?\n' \
      '1 - Минимальная цена при совпадении артикулов\n' \
      '2 - Приоритет цены по важности прайса\n\n' \
      'Введите номер варианта: '

merge_mode = input(msg)


def _get_files():
    """ Получем словарь обрабатываемых прайсов:
    ключ - имя файла, значение - путь файла """
    global files
    files = {}
    for file in os.listdir(r'.\input-for-merge'):
        path = os.path.abspath('input-for-merge\\' + file)
        files[file] = {'path': path}
        files[file]['priority'] = 0
    return files


def _get_priority():
    """ Запрашиваем приоритет цен в прайсах. Если приоритет прайса выше
        (значение больше), то цена перезаписывается. """
    msg = '\nВведите приоритет перезаписи прайсов. Если приоритет выше (значение больше),\n' \
          'то, при совпадении артикулов, цена перезапишется на цену с бОльшим значением.\n\n' \
          'Список файлов:\n'
    print(msg)
    for file in files.keys():
        print('+', file)
    print()
    for file in files.keys():
        files[file]['priority'] = int(input(f'Введите приоритет для файла "{file}": '))


def _write_data_to_file(filename, my_data):
    """ Запись обработанных данных в файл в нужном формате"""
    wb = Workbook()
    ws = wb.active
    for i, row in enumerate(my_data, start=1):
        # ws.cell(row, column, value)
        try:
            ws.cell(i, 1, row[0]).number_format = BUILTIN_FORMATS[1]  # (1: 0)        Артикул
            ws.cell(i, 2, row[1]).number_format = BUILTIN_FORMATS[0]  # (0: General)  Наименование
            ws.cell(i, 3, row[2]).number_format = BUILTIN_FORMATS[2]  # (2: 0.00)     Цена
            ws.cell(i, 4, row[3]).number_format = BUILTIN_FORMATS[2]  # (2: 0.00)     Цена со скидкой
            ws.cell(i, 5, row[4]).number_format = BUILTIN_FORMATS[0]  # (0: General)  Тип скидки
        except IndexError:
            print('IndexError : Не верное кол-во входных колонок')

    _make_table_view(ws)
    _cell_alignment(ws)
    wb.save(filename)


def merge_prices(merge_mode):
    """ Объединяет прайсы в один исходя из выбранного режима """
    data_set = {'артикул':
                    {'priority': 0,
                     'row': ['Артикул', 'Наименование', 'Цена', 'Цена со скидкой', 'Тип скидки']
                     }
                }
    for file in files:
        book_in = load_workbook(filename=files[file]['path'], data_only=True)
        sheet_in = book_in.active
        current_priority = int(files[file]['priority'])
        for current_row in sheet_in.values:
            art = str(current_row[0]).lower()
            if art not in data_set.keys():
                data_set[art] = dict(priority=current_priority, row=list(current_row))
            else:
                if merge_mode == '1' and data_set[art]['row'][2] != 'Цена' and \
                        float(current_row[2]) < float(data_set[art]['row'][2]):
                    data_set[art]['row'] = list(current_row)
                if merge_mode == '2' and current_priority > data_set[art]['priority']:
                    data_set[art]['priority'] = current_priority
                    data_set[art]['row'] = list(current_row)
    new_data = [v['row'] for v in data_set.values()]
    _write_data_to_file('merged_price.xlsx', new_data)


if __name__ == '__main__':
    _get_files()
    if merge_mode == '2':
        _get_priority()

    merge_prices(merge_mode)
