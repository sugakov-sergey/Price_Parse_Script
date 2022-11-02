"""
Объединяет все файлы из папки 'input-for-merge' в один
файл 'merged_price'. Все файлы должны быть сделаны по формату
из скрипта 'start_app.py'. При этом добавялется новая колонка
Бренд.
"""
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import BUILTIN_FORMATS
from tqdm import tqdm, trange


def get_files():
    """ Получем словарь обрабатываемых прайсов:
    ключ - имя файла, значение - путь файла """
    files_dict = {}
    try:
        for file in os.listdir(r'.\input-for-merge'):
            brand = file.split('.')[0]
            path = os.path.abspath('input-for-merge\\' + file)
            files_dict[brand] = {'path': path}
    except Exception as e:
        print(e)
    return files_dict


def merge_prices(files_dict: dict):
    """ ... """
    data_list = [['Бренд', 'Артикул', 'Наименование', 'Цена', 'Цена со скидкой', 'Тип скидки']]
    for file in tqdm(files_dict, desc='1/3 Объединение прайсов'):
        wb = load_workbook(filename=files_dict[file]['path'], data_only=True)
        ws = wb.active
        for row in ws.values:
            if row[0] != 'Артикул':
                data_list += [[file.title()] + [*row]]
    return data_list


def write_data_to_file(filename, my_data):
    """ Запись обработанных данных в файл в нужном формате"""
    wb = Workbook()
    ws = wb.active
    for i in trange(len(my_data), desc='2/3 Запись в файл      '):
        # ws.cell(row, column, value)
        try:
            ws.cell(i+1, 1, my_data[i][0]).number_format = BUILTIN_FORMATS[0]  # (0: General)  Бренд
            ws.cell(i+1, 2, my_data[i][1]).number_format = BUILTIN_FORMATS[1]  # (1: 0)        Артикул
            ws.cell(i+1, 3, my_data[i][2]).number_format = BUILTIN_FORMATS[0]  # (0: General)  Наименование
            ws.cell(i+1, 4, my_data[i][3]).number_format = BUILTIN_FORMATS[2]  # (2: 0.00)     Цена
            ws.cell(i+1, 5, my_data[i][4]).number_format = BUILTIN_FORMATS[2]  # (2: 0.00)     Цена со скидкой
            ws.cell(i+1, 6, my_data[i][5]).number_format = BUILTIN_FORMATS[0]  # (0: General)  Тип скидки
        except IndexError:
            print('IndexError : Не верное кол-во входных колонок')

    _make_table_view(ws)
    _cell_alignment(ws)
    wb.save(filename)


def _make_table_view(sheet):
    """ Настраиваем таблицу для вывода данных """
    # Задаем ширину колонки
    sheet.column_dimensions['A'].width = 20  # Бренд
    sheet.column_dimensions['B'].width = 30  # Артикул
    sheet.column_dimensions['C'].width = 60  # Наименование
    sheet.column_dimensions['D'].width = 15  # Цена
    sheet.column_dimensions['E'].width = 20  # Цена со скидкой
    sheet.column_dimensions['F'].width = 20  # Тип скидки
    # Фиксируем строку заголовка
    sheet.freeze_panes = 'A2'


def _cell_alignment(sheet):
    """ Выравнивание данных в ячейках по центру """
    for col in trange(1, 7, desc='3/3 Выравнивание данных'):
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')


if __name__ == '__main__':
    files = get_files()
    data = merge_prices(files)
    write_data_to_file('merged_by_brand.xlsx', data)
    print('Файл "merged_by_brand.xlsx" готов! ')
