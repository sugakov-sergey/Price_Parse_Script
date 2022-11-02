"""
Переводит артикулы Laufen и Jika из формата БЕЗ ТОЧЕК
в формат С ТОЧКАМИ. Сначала сверяет артикул с существующим прайсом.
Если такого артикула не находит, то переводит запись в
формат X.XXXX.X.XXX.XXX.X
"""
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from tqdm import tqdm, trange


def get_filepath_without_dot():
    """..."""
    res, files = '', {}
    for file in os.listdir(r'.\price_without_dots'):
        path = os.path.abspath('price_without_dots\\' + file)
        files[path] = os.stat(path).st_mtime
    res = max(files, key=files.get)
    return res


def get_filepath_with_dot():
    """..."""
    res, files = '', {}
    for file in os.listdir(r'.\general_price_with_dots'):
        path = os.path.abspath('general_price_with_dots\\' + file)
        files[path] = os.stat(path).st_mtime
    res = max(files, key=files.get)
    return res


def make_artikul_set_with_dot():
    """ ... """
    data_set = set()
    price_with_dots = load_workbook(filename=get_filepath_with_dot(), data_only=True)
    ws = price_with_dots.active
    for row in ws.values:
        if row[0] != 'Артикул':
            data_set.add(row[0])
    return data_set


def make_data_with_dots():
    """ ... """
    new_data = [['Артикул', 'Наименование', 'Цена', 'Цена со скидкой', 'Тип скидки']]
    price_without_dot = load_workbook(filename=get_filepath_without_dot(), data_only=True)
    ws = price_without_dot.active
    for row in ws.values:
        if row[0] != 'Артикул':
            new_data += [[*row]]
    return new_data


def make_dots(my_data: list, da_set: set):
    """..."""
    data_with_dots = []
    for row in my_data:
        data_with_dots += [check_and_replace(row, da_set)]
    return data_with_dots


def check_and_replace(row: list, d_set: set):
    """..."""
    if row[0] and row[0] != 'Артикул':
        for art in d_set:
            if str(row[0]) == str(''.join(art.split('.'))):
                row[0] = art
                return row
        art = str(row[0])
        new_art = f'{art[0]}.{art[1:5]}.{art[5]}.{art[6:9]}.{art[9:12]}.{art[12]}'
        row[0] = new_art
        return row


def write_data_to_file(filename, my_data):
    """ Запись обработанных данных в файл """
    book = Workbook()
    sheet = book.active
    _make_table_view(sheet)

    for row in my_data:
        sheet.append(row)

    _cell_alignment(sheet)
    book.save(filename)


def _make_table_view(sheet):
    """ Настраиваем таблицу для вывода данных """
    # Задаем ширину колонки
    sheet.column_dimensions['A'].width = 30  # Артикул
    sheet.column_dimensions['B'].width = 60  # Наименование
    sheet.column_dimensions['C'].width = 15  # Цена
    sheet.column_dimensions['D'].width = 20  # Цена со скидкой
    sheet.column_dimensions['E'].width = 20  # Тип скидки
    # Фиксируем строку заголовка
    sheet.freeze_panes = 'A2'


def _cell_alignment(sheet):
    """ Выравнивание данных в ячейках по центру """
    for col in trange(1, 7, desc='3/3 Выравнивание данных'):
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')


if __name__ == '__main__':
    get_filepath_without_dot()
    data_set = make_artikul_set_with_dot()
    my_data = make_data_with_dots()
    new_data = make_dots(my_data, data_set)
    write_data_to_file('jika-byn.xlsx', my_data)


