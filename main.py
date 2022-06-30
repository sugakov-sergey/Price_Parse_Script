import json
import os
from random import randint
from openpyxl import load_workbook, Workbook
from itertools import chain
from openpyxl.styles import Alignment

import config

data, line, not_valid_values = [], [], []
recurring, brands, brand_config = {}, {}, {}
price_is_valid, art_is_valid = False, False
brand = ''


def clear_temp_files():
    """ Очистка временных файлов в начале нового цикла """
    if os.path.exists('log.txt'):
        os.remove('log.txt')


def input_brand():
    """ Получаем название бренда, по которому применяются настройки """
    global brand
    in_brand = input("Введите название брэнда: ")
    brand = in_brand.strip().lower().replace(' ', '').replace('-', '')
    _get_config(brand)


def _get_config(brand):
    """ Получаем конфигурацию прайса по названию бренда """
    global brands, brand_config
    with open('brands_config.json', 'r') as file:
        brands = json.load(file)

    if brand not in brands:
        _add_to_config()
    else:
        brand_config = brands[brand]


def _add_to_config():
    """ Добавляем конфигурацию колонок в файл для будущего использования """
    _print_sheet_names(book_in)
    brand_config['sheet_number'] = int(input('\nВведите номер Листа: ')) - 1
    brand_config['art'] = int(input('Введите номер колонки Арикула: '))
    brand_config['title'] = int(input('Введите номер колонки Наименования: '))
    brand_config['price'] = int(input('Введите номер колонки Цена: '))
    brand_config['price_dis'] = int(input('Введите номер колонки Цена со скидкой: '))
    brands[brand] = brand_config
    with open('brands_config.json', 'w') as file:
        json.dump(brands, file)


def load_price():
    """ Загружаем входящий прайс для обработки.
        Прайс должен находиться в папке input """
    global sheet_in, book_in
    book_in = load_workbook(filename=_get_filename(), data_only=True)
    # Укажите актуальный лист
    input_brand()
    sheet_in = book_in[book_in.sheetnames[brands[brand]['sheet_number']]]


def _print_sheet_names(book_in):
    """ Выводим на экран все листы книги с их индексами """
    print('\n', '-' * 12, 'Список листов с индексами', '-' * 13, '\n')
    for i, sheet in enumerate(book_in.sheetnames, 1):
        print(i, sheet)


def _get_filename():
    """ Возвращает путь файла, который был добавлен в папку последним """
    files = {}
    for file in os.listdir(r'.\input'):
        path = os.path.abspath('input\\' + file)
        files[path] = os.stat(path).st_mtime
    return max(files, key=files.get)


def info():
    """ Информация о данных в переданном листе """
    # Структура данных
    print('\n', '-' * 12, 'Структура исходных данных', '-' * 13, '\n')
    for row in sheet_in.iter_rows(min_row=0, max_row=6, max_col=6, values_only=True):
        print(row)
    print('...')
    print(sheet_in.max_row, 'строк')

    # Список невалидных строк
    if not_valid_values:
        print('\n', '-' * 10, 'Список строк с неверными данными ', '-' * 10, '\n')
        for i in range(min(5, len(not_valid_values))):
            print(not_valid_values[randint(0, len(not_valid_values) - 1)])
        print('...')
        print(len(not_valid_values), 'строк')

    # Список дублированных строк
    if sum(recurring.values()) > 0:
        print('\n', '-' * 14, 'Строки с дублированными артикулами', '-' * 14, '\n')
        for art, count in recurring.items():
            if count > 0:
                print(art, '...', count, 'дублей')
        print('...')
        print(sum(recurring.values()), 'строк')

    # Сравнение строк до/после
    print('\n', '-' * 14, 'Сравнение строк до/после', '-' * 14, '\n')
    print(sheet_in.max_row, 'было -', len(not_valid_values), 'невалидных -',
          sum(recurring.values()), 'дублей =', len(data), 'стало')
    before = (sheet_in.max_row - len(not_valid_values) - sum(recurring.values()))
    after = len(data)
    print(before == after)


def make_book(art, title, price, price_dis, type_dis):
    """ Создаем новый список строк значений в нужном порядке """
    global data, line
    cols = [art, title, price, price_dis, type_dis]
    data = [[sheet_in.cell(row, col).value for col in cols] for row in range(1, sheet_in.max_row + 1)]


def clear_data(my_data):
    for row in my_data[:]:
        _clear_art_cell(row)
        _clear_price_cell(row)
        _clear_price_dis_cell(row)
        _remove_invalid_values(row)


def _log_this(row, reason):
    with open('log.txt', 'a') as f:
        f.writelines([str(row), '\n', reason, '\n', '\n'])


def _remove_invalid_values(row):
    global not_valid_values
    global recurring
    if not _is_valid():
        not_valid_values += [row]
        _log_this(row, 'Invalid values')
        data.remove(row)
    elif row[0] in recurring.keys():
        recurring[row[0]] += 1
        _log_this(row, 'Recurring values')
        data.remove(row)
    else:
        recurring[row[0]] = 0


def _clear_art_cell(row):
    """ Форматирует ячейку с артикулом"""
    global art_is_valid
    art_is_valid = False
    if isinstance(row[0], str):
        row[0] = row[0].strip()
        art_is_valid = True
    elif isinstance(row[0], int | float):
        row[0] = int(row[0])
        art_is_valid = True


def _clear_price_cell(row):
    """ Форматирует ячейку цены в число с двумя знаками после запятой"""
    global price_is_valid
    price_is_valid = False
    if isinstance(row[2], str):
        if set(row[2]) <= set('0123456789., '):
            row[2] = row[2].strip().replace(',', '.').replace(' ', '')
            row[2] = round(float(row[2]), 2)
            if row[2] > 0:
                price_is_valid = True
    elif isinstance(row[2], float | int) and row[2] > 0:
        row[2] = round(float(row[2]), 2)
        price_is_valid = True


def _clear_price_dis_cell(row):
    """ Форматирует ячейку цены со скидкой в число с двумя знаками после запятой.
        А так же при наличии скидки, прописывает тип скидки """
    if row[3]:
        if isinstance(row[3], str):
            if set(row[3]) <= set('0123456789., '):
                row[3] = row[3].strip().replace(',', '.').replace(' ', '')
                row[3] = round(float(row[3]), 2)
        elif isinstance(row[3], float | int):
            row[3] = round(float(row[3]), 2)
        row[4] = 'Акция'


def _is_valid():
    """ Валидация данных в необходимых для загрузки ячейках """
    return all((art_is_valid, price_is_valid))


def write_data_to_file(filename, my_data):
    """ Запись обработанных данных в файл """
    book = Workbook()
    sheet = book.active
    _make_table_view(sheet)
    _sort_by_discount(my_data)

    for row in chain(config.title_table, my_data):
        sheet.append(row)

    _cell_alignment(sheet)
    book.save(filename)


def _sort_by_discount(my_data):
    """ Сортирует данные по колонке 'Цена со скидкой' """

    def none_sorter(_data):
        if not _data[3]:
            return 0
        return _data[3]

    my_data.sort(key=none_sorter, reverse=True)


def _cell_alignment(sheet):
    for col in range(1, 6):
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')


def _make_table_view(sheet):
    """ Настраиваем таблицу для вывода данных """
    # изменяем ширину колонки
    sheet.column_dimensions['A'].width = 30  # Артикул
    sheet.column_dimensions['B'].width = 60  # Наименование
    sheet.column_dimensions['C'].width = 15  # Цена
    sheet.column_dimensions['D'].width = 20  # Цена со скидкой
    sheet.column_dimensions['E'].width = 20  # Тип скидки
    # Фиксируем строку заголовка
    sheet.freeze_panes = 'A2'


clear_temp_files()
load_price()
make_book(brands[brand]['art'], brands[brand]['title'],
          brands[brand]['price'], brands[brand]['price_dis'], 100)
clear_data(data)
write_data_to_file(config.filename_out, data)
info()
