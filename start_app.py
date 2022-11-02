import json
import os
import shutil
import sys
from random import randint
import openpyxl
from openpyxl import load_workbook, Workbook
from itertools import chain
from openpyxl.styles import Alignment
from tqdm import trange
from transliterate import translit

from config import config
from modules.log import logging as log

data, line, not_valid_values = [], [], []
recurring, brands, brand_config = {}, {}, {}
price_is_valid, art_is_valid = False, False
brand, currency, brand_price_path = '', '', ''
current_row, encode_error_values = [], []


def clear_temp_files():
    """ Очистка временных файлов в начале нового цикла """
    log.info('start')
    if os.path.exists('logs/invalid_values.log'):
        os.remove('logs/invalid_values.log')
    log.info('finish')


def input_brand():
    """ Получаем название бренда, по которому применяются настройки """
    log.info('start')
    global brand
    in_brand = input("Введите название брэнда (EN/RU): ")
    brand = in_brand.strip().lower().replace(' ', '').replace('-', '')
    brand = translit(brand, language_code='ru', reversed=True)
    _get_config()
    log.info(f'finish\n{brand}')


def _get_config():
    """ Получаем конфигурацию прайса по названию бренда """
    log.info('start')
    global brand, brands, brand_config
    with open('config/brands_config.json', 'r') as file:
        brands = json.load(file)
    if brand not in brands:
        _add_to_config()
    elif len(brands[brand]) < 6:
        _add_to_config()
    else:
        brand_config = brands[brand]
    log.info(f'finish\n{brands[brand]}')


def _add_to_config():
    """ Добавляем конфигурацию колонок в файл для будущего использования """
    log.info('start')
    _print_sheet_names(book_in)
    brand_config['sheet_number'] = int(input('\nВведите номер Листа: ')) - 1
    brand_config['art'] = int(input('Введите номер колонки Арикула: '))
    brand_config['title'] = int(input('Введите номер колонки Наименования: '))
    brand_config['price'] = int(input('Введите номер колонки Цена: '))
    brand_config['price_dis'] = int(input('Введите номер колонки Цена со скидкой: '))
    brand_config['currency'] = input('Введите Валюту прайса (byn/usd/eur): ').lower()
    brands[brand] = brand_config
    with open('config/brands_config.json', 'w') as file:
        json.dump(brands, file)
    log.info(f'finish\n{brand_config}')


def load_price():
    """ Загружаем входящий прайс для обработки. Прайс должен находиться в папке input """
    log.info('start')
    global sheet_in, book_in
    try:
        book_in = load_workbook(filename=_get_filename(), data_only=True)
        # Укажите актуальный лист
        input_brand()
        sheet_in = book_in[book_in.sheetnames[brands[brand]['sheet_number']]]
        log.info('finish')
    except openpyxl.utils.exceptions.InvalidFileException:
        sys.exit("Format error. Please save price in .xlsx format and Try Again")


def _print_sheet_names(book):
    log.info('start')
    """ Выводим на экран все листы книги с их индексами """
    print('\n', '-' * 12, 'Список листов с индексами', '-' * 13, '\n')
    for i, sheet in enumerate(book.sheetnames, 1):
        print(i, sheet)
    log.info('finish')


def _get_filename():
    """ Возвращает путь файла, который был добавлен в папку последним """
    log.info('start')
    res, files = '', {}
    for file in os.listdir(r'.\input'):
        path = os.path.abspath('input\\' + file)
        files[path] = os.stat(path).st_mtime
    res = max(files, key=files.get)
    log.info(f'finish\n{res}')
    return res


def info():
    """ Информация о данных в переданном листе """
    log.info('start')
    # Структура данных
    print('\n', '-' * 12, 'Структура исходных данных', '-' * 13, '\n')
    for row in sheet_in.iter_rows(min_row=0, max_row=6, max_col=6, values_only=True):
        try:
            print(row)
        except UnicodeEncodeError:
            print('UnicodeEncodeError')
    print('...')
    print(sheet_in.max_row, 'строк')

    # Список невалидных строк
    if not_valid_values:
        print('\n', '-' * 10, 'Список строк с неверными данными ', '-' * 10, '\n')
        for i in range(min(5, len(not_valid_values))):
            try:
                print(not_valid_values[randint(0, len(not_valid_values) - 1)])
            except UnicodeEncodeError:
                print('UnicodeEncodeError')
        print('...')
        print(len(not_valid_values), 'строк')

    # Список строк с неверной кодировкой
    if encode_error_values:
        print('\n', '-' * 10, 'Список строк с неверной кодировкой ', '-' * 10, '\n')
        for i in range(min(5, len(encode_error_values))):
            print(encode_error_values[randint(0, len(encode_error_values) - 1)])
        print('...')
        print(len(encode_error_values), 'строк')

    # Список дублированных строк
    if sum(recurring.values()) > 0:
        print('\n', '-' * 14, 'Строки с дублированными артикулами', '-' * 14, '\n')
        for art, count in recurring.items():
            if count > 0:
                print(art, '...', count, 'дублей')
        print('...')
        print(sum(recurring.values()), 'строк')

    # Сравнение строк до/после
    print('\n', '-' * 14, 'Сравнение строк до/после', '-' * 14, '\n')
    print(sheet_in.max_row, 'было -', len(not_valid_values), 'невалидных -',
          sum(recurring.values()), 'дублей -', len(encode_error_values),
          'неверная кодировка =', len(data), 'стало')
    before = (sheet_in.max_row - len(not_valid_values) -
              len(encode_error_values) - sum(recurring.values()))
    after = len(data)
    print(before == after)
    log.info('finish')


def make_book(art, title, price, price_dis, type_dis):
    """ Создаем новый список строк значений в нужном порядке """
    log.info('start')
    global data, line
    cols = [art, title, price, price_dis, type_dis]
    data = [[sheet_in.cell(row, col).value for col in cols] for row in range(1, sheet_in.max_row + 1)]
    log.info('finish')


def clear_data(my_data):
    """ Чистим построчно каждую ячейку перед записью """
    # progressbar для визуализации работы больших файлов
    log.info('start')
    global current_row, encode_error_values
    my_data_copy = my_data[:]
    for i in trange(len(my_data_copy), colour='yellow', desc='Обработка прайса'):
        row = my_data_copy[i]
        try:
            _clear_art_cell(row)
            _clear_price_cell(row)
            _clear_price_dis_cell(row)
            _remove_invalid_values(row)
            current_row = row
        except UnicodeEncodeError:
            current_row = [i.encode('utf-8', 'ignore') for i in row if isinstance(i, str)]
            encode_error_values += [current_row]
            _log_invalid(current_row, 'UnicodeEncodeError values')
            data.remove(row)
            not_valid_values.pop()
    log.info(f'finish\nlen(my_data) = {len(my_data)}')


def _log_invalid(row: list, reason: str):
    """ Ведем логи не вошедших строк с указанием причины """
    log.info(f'start\n{row}')
    with open('logs/invalid_values.log', 'a') as f:
        f.writelines([str(row), '\n', reason, '\n', '\n'])
    log.info('finish')


def _remove_invalid_values(row: list):
    """ Проверяем данные и удаляем невалидные """
    log.info(f'start\n{row}')
    global not_valid_values
    global recurring
    if not _is_valid():
        not_valid_values += [row]
        _log_invalid(row, 'Invalid values')
        data.remove(row)
    elif row[0] in recurring.keys():
        recurring[row[0]] += 1
        _log_invalid(row, 'Recurring values')
        data.remove(row)
    else:
        recurring[row[0]] = 0
    log.info('finish')


def _clear_art_cell(row: list):
    """ Форматирует ячейку с артикулом"""
    log.info(f'start\n{row[0]}')
    global art_is_valid
    art_is_valid = False
    if isinstance(row[0], str):
        row[0] = ' '.join(row[0].split())  # удаляем лишние пробелы
        art_is_valid = True
    elif isinstance(row[0], int | float):
        row[0] = int(row[0])
        art_is_valid = True
    log.info(f'finish\n{row[0]}')


def _clear_price_cell(row: list):
    """ Форматирует ячейку цены в число с двумя знаками после запятой"""
    log.info(f'start\n{row[2]}')
    global price_is_valid
    price_is_valid = False
    if isinstance(row[2], str):
        if set(row[2]) <= set('0123456789., '):
            row[2] = row[2].strip().replace(',', '.').replace(' ', '')
            row[2] = round(float(row[2]), 2)
            if row[2] > 0:
                price_is_valid = True
    elif isinstance(row[2], float | int) and row[2] > 0:
        row[2] = round(float(row[2]), 2)
        price_is_valid = True
    log.info(f'finish\n{row[2]}')


def _clear_price_dis_cell(row: list):
    """ Форматирует ячейку цены со скидкой в число с двумя знаками после запятой.
        А так же при наличии скидки, прописывает тип скидки """
    log.info(f'start\n{row[3]}')
    if row[3]:
        if isinstance(row[3], str):
            if set(row[3]) <= set('0123456789., '):
                row[3] = row[3].strip().replace(',', '.').replace(' ', '')
                row[3] = round(float(row[3]), 2)
        elif isinstance(row[3], float | int):
            row[3] = round(float(row[3]), 2)
        row[4] = 'Акция'
    log.info(f'finish\n{row[3]}')


def _is_valid():
    """ Валидация данных в необходимых для загрузки ячейках """
    log.info('start')
    res = all((art_is_valid, price_is_valid))
    log.info(f'{res}')
    log.info('finish')
    return res


def write_data_to_file(filename, my_data):
    """ Запись обработанных данных в файл """
    log.info(f'start\nlen(my_data) = {len(my_data)}')
    global brand_price_path
    book = Workbook()
    sheet = book.active
    _make_table_view(sheet)
    _sort_by_discount(my_data)

    for row in chain(config.title_table, my_data):
        sheet.append(row)

    _cell_alignment(sheet)
    book.save(filename)

    # Сохраняем копию в папку output
    with open('config/brands_config.json', 'r') as file:
        brands = json.load(file)
    brand_price_path = f'output\\{brand}-{brands[brand]["currency"]}.xlsx'
    shutil.copy(filename, brand_price_path)
    log.info('finish')


def _sort_by_discount(my_data):
    """ Сортирует данные по колонке 'Цена со скидкой' """
    log.info(f'start\nlen(my_data) = {len(my_data)}')

    def none_sorter(_data):
        if not _data[3]:
            return 0
        return _data[3]

    my_data.sort(key=none_sorter, reverse=True)
    log.info('finish')


def _cell_alignment(sheet):
    """ Выравнивание данных в ячейках по центру """
    log.info('start')
    for col in range(1, 6):
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    log.info('finish')


def _make_table_view(sheet):
    """ Настраиваем таблицу для вывода данных """
    # Задаем ширину колонки
    log.info('start')
    sheet.column_dimensions['A'].width = 30  # Артикул
    sheet.column_dimensions['B'].width = 60  # Наименование
    sheet.column_dimensions['C'].width = 15  # Цена
    sheet.column_dimensions['D'].width = 20  # Цена со скидкой
    sheet.column_dimensions['E'].width = 20  # Тип скидки
    # Фиксируем строку заголовка
    sheet.freeze_panes = 'A2'
    log.info('finish')


if __name__ == "__main__":
    clear_temp_files()
    load_price()
    make_book(brands[brand]['art'], brands[brand]['title'],
              brands[brand]['price'], brands[brand]['price_dis'], 100)
    clear_data(data)
    write_data_to_file(config.filename_out, data)
    info()
